"""
Excel exporter — generates a professional DCF workbook with 5 sheets:
  1. Dashboard  — KPIs + recommendation summary
  2. DCF Model  — dynamic model with formula-driven valuation
  3. Comparables — trading comps and transaction comps
  4. Sensitivity — WACC × Terminal Growth equity value matrix
  5. Financials  — income statement
"""
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import AnalysisResult

# ── Palette ────────────────────────────────────────────────────────────────
NAVY = "1B2A4A"
DARK_BLUE = "2C3E50"
GOLD = "C5A028"
WHITE = "FFFFFF"
LIGHT_BLUE = "EBF0F7"
LIGHT_YELLOW = "FFFDE7"
DARK_GRAY = "333333"
MID_GRAY = "888888"
GREEN_BG = "E8F5E9"
RED_BG = "FFEBEE"
GREEN_FONT = "1B5E20"
RED_FONT = "B71C1C"
STRIPE = "F7F9FC"


# ── Style helpers ───────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _bottom_border() -> Border:
    b = Side(style="thin", color="AAAAAA")
    return Border(bottom=b)


def _font(size: int = 10, bold: bool = False, color: str = DARK_GRAY, italic: bool = False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left(indent: int = 1) -> Alignment:
    return Alignment(horizontal="left", vertical="center", indent=indent)


def _style_title_row(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.fill = _fill(NAVY)
    c.font = _font(14, bold=True, color=WHITE)
    c.alignment = _center()
    ws.row_dimensions[row].height = 34


def _style_section(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.fill = _fill(DARK_BLUE)
    c.font = _font(10, bold=True, color=WHITE)
    c.alignment = _left()
    ws.row_dimensions[row].height = 20


def _style_col_headers(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i)
        c.value = h
        c.fill = _fill("34495E")
        c.font = _font(9, bold=True, color=WHITE)
        c.alignment = _center()
        c.border = _thin_border()
    ws.row_dimensions[row].height = 18


def _data_cell(ws, row: int, col: int, value, fmt: str = "", bold: bool = False,
               bg: str = "", align: str = "center") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(10, bold=bold)
    c.border = _thin_border()
    c.alignment = Alignment(horizontal=align, vertical="center", indent=(1 if align == "left" else 0))
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = _fill(bg)


# ── Value parsers ───────────────────────────────────────────────────────────

def _parse_num(s: Optional[str]) -> Optional[float]:
    """Parse '$4.2B', '$800M', '1000' → raw float (not millions)."""
    if not s:
        return None
    s = str(s).strip().replace(",", "").replace("+", "")
    upper = s.upper()
    mult = 1e9 if "B" in upper else (1e6 if "M" in upper else (1e3 if "K" in upper else 1.0))
    is_pct = "%" in s
    clean = re.sub(r"[^0-9.\-]", "", s)
    try:
        val = float(clean) * mult
        return val / 100 if is_pct else val
    except ValueError:
        return None


def _parse_pct(s: Optional[str]) -> Optional[float]:
    """Parse '18%', '+22%' → 0.18."""
    if not s:
        return None
    clean = re.sub(r"[^0-9.\-]", "", str(s).replace("+", ""))
    try:
        return float(clean) / 100
    except ValueError:
        return None


def _to_m(v: Optional[float]) -> Optional[float]:
    """Convert raw value to $M."""
    return v / 1e6 if v is not None else None


# ── Sheet builders ──────────────────────────────────────────────────────────

def _build_dashboard(wb: Workbook, r: AnalysisResult) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # Column widths
    for col, w in zip("ABCDE", [28, 22, 22, 22, 20]):
        ws.column_dimensions[col].width = w

    f, m, fin, v = r.fundamentals, r.market, r.financials, r.valuation

    _style_title_row(ws, 1, f"  {f.company_name}  —  Gold Roger Dashboard", 5)

    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}   |   Sector: {f.sector or 'N/A'}   |   {f.headquarters or ''}"
    ws["A2"].font = _font(9, color=MID_GRAY, italic=True)
    ws["A2"].alignment = _left()
    ws.merge_cells("A2:E2")

    # Recommendation badge
    _style_section(ws, 4, "RECOMMENDATION", 5)
    rec = (v.recommendation or "N/A").upper()
    bg = GREEN_BG if rec == "BUY" else (RED_BG if rec == "SELL" else LIGHT_YELLOW)
    fg = GREEN_FONT if rec == "BUY" else (RED_FONT if rec == "SELL" else DARK_BLUE)

    for col, (lbl, val) in enumerate(
        [
            ("Recommendation", rec),
            ("Target Value", v.implied_value or "N/A"),
            ("Upside / Downside", v.upside_downside or "N/A"),
            ("Current Price", v.current_price or "N/A"),
        ],
        start=1,
    ):
        ws.cell(row=5, column=col).value = lbl
        ws.cell(row=5, column=col).font = _font(9, bold=True, color=MID_GRAY)
        ws.cell(row=5, column=col).alignment = _center()
        ws.cell(row=6, column=col).value = val
        ws.cell(row=6, column=col).font = _font(14, bold=True, color=fg)
        ws.cell(row=6, column=col).fill = _fill(bg)
        ws.cell(row=6, column=col).alignment = _center()
        ws.row_dimensions[6].height = 32

    # KPI table
    _style_section(ws, 8, "KEY FINANCIALS", 5)
    _style_col_headers(ws, 9, ["Metric", "Value", "Metric", "Value"], 1)
    kpis = [
        ("Revenue", fin.revenue_current),
        ("Revenue Growth", fin.revenue_growth),
        ("Gross Margin", fin.gross_margin),
        ("EBITDA Margin", fin.ebitda_margin),
        ("Net Margin", fin.net_margin),
        ("Free Cash Flow", fin.free_cash_flow),
        ("TAM", m.market_size),
        ("Market CAGR", m.market_growth),
    ]
    for i in range(0, len(kpis), 2):
        row = 10 + i // 2
        bg = STRIPE if (i // 2) % 2 == 0 else WHITE
        for j in range(2):
            idx = i + j
            if idx < len(kpis):
                label, val = kpis[idx]
                _data_cell(ws, row, j * 2 + 1, label, bold=True, bg=bg, align="left")
                _data_cell(ws, row, j * 2 + 2, val or "N/A", bg=bg)

    # Market context
    _style_section(ws, 15, "MARKET CONTEXT", 5)
    _style_col_headers(ws, 16, ["", "Detail"], 1)
    market_rows = [
        ("Market Segment", m.market_segment),
        ("Company Share", m.company_market_share),
        ("Top Competitor", m.main_competitors[0].name if m.main_competitors else "N/A"),
    ]
    for i, (lbl, val) in enumerate(market_rows):
        bg = STRIPE if i % 2 == 0 else WHITE
        _data_cell(ws, 17 + i, 1, lbl, bold=True, bg=bg, align="left")
        _data_cell(ws, 17 + i, 2, val or "N/A", bg=bg)
        ws.merge_cells(f"B{17+i}:E{17+i}")

    # Valuation football field
    if v.methods:
        _style_section(ws, 21, "VALUATION FOOTBALL FIELD  ($)", 5)
        _style_col_headers(ws, 22, ["Method", "Low", "Mid", "High", "Weight"], 1)
        for i, meth in enumerate(v.methods):
            bg = STRIPE if i % 2 == 0 else WHITE
            _data_cell(ws, 23 + i, 1, meth.name, bold=True, bg=bg, align="left")
            _data_cell(ws, 23 + i, 2, meth.low or "—", bg=bg)
            _data_cell(ws, 23 + i, 3, meth.mid or "—", bg=bg)
            _data_cell(ws, 23 + i, 4, meth.high or "—", bg=bg)
            _data_cell(ws, 23 + i, 5, f"{meth.weight}%" if meth.weight else "—", bg=bg)


def _build_dcf_sheet(wb: Workbook, r: AnalysisResult) -> None:
    ws = wb.create_sheet("DCF Model")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGHI", [32, 16, 14, 14, 14, 14, 14, 16, 14]):
        ws.column_dimensions[col].width = w

    f, fin, v = r.fundamentals, r.financials, r.valuation

    _style_title_row(ws, 1, f"  {f.company_name}  —  DCF Valuation Model  (dynamic)", 8)
    ws["A2"].value = f"Generated: {datetime.now().strftime('%B %d, %Y')}   |   All values in $M"
    ws["A2"].font = _font(9, color=MID_GRAY, italic=True)
    ws.merge_cells("A2:H2")

    # ── Assumptions (rows 4-8) ─────────────────────────────────────────────
    _style_section(ws, 4, "ASSUMPTIONS  (edit yellow cells to recalculate)", 3)

    dcf = v.dcf_assumptions
    wacc_val = _parse_pct(dcf.wacc if dcf else None) or 0.10
    tgr_val = _parse_pct(dcf.terminal_growth if dcf else None) or 0.03

    cur_rev_raw = _parse_num(fin.revenue_current)
    net_debt_m = 0.0  # simplified; no net debt field in model

    assumptions = [
        ("WACC", wacc_val, "0.0%", "Weighted Average Cost of Capital"),
        ("Terminal Growth Rate", tgr_val, "0.0%", "Gordon Growth terminal rate"),
        ("FCF / EBITDA Conversion", 0.70, "0%", "Assumed FCF = 70% of EBITDA"),
        ("Net Debt ($M)", net_debt_m, "#,##0.0", "Total debt minus cash"),
    ]
    # Row 5 → B5 = WACC, Row 6 → B6 = TGR, Row 7 → B7 = FCF conv, Row 8 → B8 = Net Debt
    for i, (label, value, fmt, note) in enumerate(assumptions, start=5):
        ws[f"A{i}"].value = label
        ws[f"A{i}"].font = _font(10)
        ws[f"A{i}"].alignment = _left()
        ws[f"B{i}"].value = value
        ws[f"B{i}"].number_format = fmt
        ws[f"B{i}"].font = _font(10, bold=True)
        ws[f"B{i}"].fill = _fill(LIGHT_YELLOW)
        ws[f"B{i}"].border = _thin_border()
        ws[f"C{i}"].value = note
        ws[f"C{i}"].font = _font(9, color=MID_GRAY, italic=True)
        ws[f"C{i}"].alignment = _left()
        ws.row_dimensions[i].height = 18

    # ── Projections (headers row 10, data rows 11-15) ──────────────────────
    PROJ_HEADER_ROW = 10
    PROJ_DATA_START = 11  # Year 1 → row 11, Year 5 → row 15

    _style_section(ws, PROJ_HEADER_ROW - 1, "REVENUE & CASH FLOW PROJECTIONS", 8)
    _style_col_headers(
        ws,
        PROJ_HEADER_ROW,
        ["Year", "Revenue ($M)", "Growth %", "EBITDA Margin", "EBITDA ($M)", "FCF ($M)", "Disc. Factor", "PV of FCF ($M)"],
    )

    # Parse projection data from analysis
    cur_year = datetime.now().year
    ebitda_pct = _parse_pct(fin.ebitda_margin) or 0.15
    cur_rev_m = _to_m(cur_rev_raw) or 1_000.0

    proj_rows: list[dict] = []
    for i in range(5):
        year = cur_year + i
        p = fin.projections[i] if i < len(fin.projections) else None
        rev_m = _to_m(_parse_num(p.revenue if p else None))
        growth = _parse_pct(p.growth if p else None)
        ebitda_m_pct = _parse_pct(p.ebitda_margin if p else None) or ebitda_pct

        if rev_m is None:
            if proj_rows:
                rev_m = proj_rows[-1]["revenue"] * (1 + (growth or 0.10))
            else:
                rev_m = cur_rev_m * (1 + (growth or 0.12))

        proj_rows.append({"year": year, "revenue": rev_m, "growth": growth, "ebitda_m": ebitda_m_pct})

    for i, pd in enumerate(proj_rows):
        row = PROJ_DATA_START + i
        n = i + 1  # year number (1-5)
        bg = STRIPE if i % 2 == 0 else WHITE

        _data_cell(ws, row, 1, pd["year"], bold=True, bg=bg)
        _data_cell(ws, row, 2, pd["revenue"], fmt="#,##0.0", bg=bg)
        if pd["growth"] is not None:
            _data_cell(ws, row, 3, pd["growth"], fmt="0.0%", bg=bg)
        else:
            _data_cell(ws, row, 3, "—", bg=bg)
        _data_cell(ws, row, 4, pd["ebitda_m"], fmt="0.0%", bg=bg)
        # Formulas referencing this row's Revenue and EBITDA margin
        ws.cell(row=row, column=5).value = f"=B{row}*D{row}"
        ws.cell(row=row, column=5).number_format = "#,##0.0"
        ws.cell(row=row, column=5).fill = _fill(bg)
        ws.cell(row=row, column=5).border = _thin_border()

        ws.cell(row=row, column=6).value = f"=E{row}*$B$7"
        ws.cell(row=row, column=6).number_format = "#,##0.0"
        ws.cell(row=row, column=6).fill = _fill(bg)
        ws.cell(row=row, column=6).border = _thin_border()

        ws.cell(row=row, column=7).value = f"=1/(1+$B$5)^{n}"
        ws.cell(row=row, column=7).number_format = "0.0000"
        ws.cell(row=row, column=7).fill = _fill(bg)
        ws.cell(row=row, column=7).border = _thin_border()

        ws.cell(row=row, column=8).value = f"=F{row}*G{row}"
        ws.cell(row=row, column=8).number_format = "#,##0.0"
        ws.cell(row=row, column=8).fill = _fill(bg)
        ws.cell(row=row, column=8).border = _thin_border()

    LAST_PROJ_ROW = PROJ_DATA_START + 4  # row 15

    # ── DCF Valuation Summary (rows 18-24) ────────────────────────────────
    VAL_HEADER = 17
    VAL_START = 18  # first item row

    _style_section(ws, VAL_HEADER, "DCF VALUATION SUMMARY", 8)

    row_pv_fcfs = VAL_START        # 18
    row_tv = VAL_START + 1         # 19
    row_pv_tv = VAL_START + 2      # 20
    row_ev = VAL_START + 3         # 21
    row_nd = VAL_START + 4         # 22
    row_equity = VAL_START + 5     # 23

    val_items = [
        (row_pv_fcfs, "Sum of PV FCFs ($M)", f"=SUM(H{PROJ_DATA_START}:H{LAST_PROJ_ROW})"),
        (row_tv, "Terminal Value ($M)", f"=F{LAST_PROJ_ROW}*(1+$B$6)/MAX($B$5-$B$6,0.001)"),
        (row_pv_tv, "PV of Terminal Value ($M)", f"=B{row_tv}/(1+$B$5)^5"),
        (row_ev, "Enterprise Value ($M)", f"=B{row_pv_fcfs}+B{row_pv_tv}"),
        (row_nd, "Less: Net Debt ($M)", "=$B$8"),
        (row_equity, "Equity Value ($M)", f"=B{row_ev}-B{row_nd}"),
    ]

    for row, label, formula in val_items:
        is_final = "Equity" in label
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = _font(10, bold=is_final)
        ws[f"A{row}"].alignment = _left()
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].number_format = "#,##0.0"
        ws[f"B{row}"].font = _font(11 if is_final else 10, bold=is_final)
        ws[f"B{row}"].border = _thin_border()
        if is_final:
            ws[f"A{row}"].fill = _fill(GREEN_BG)
            ws[f"B{row}"].fill = _fill(GREEN_BG)
        ws.row_dimensions[row].height = 18

    # separator line before equity
    ws.cell(row=row_equity - 1, column=1).border = Border(bottom=Side(style="medium", color=NAVY))
    ws.cell(row=row_equity - 1, column=2).border = Border(bottom=Side(style="medium", color=NAVY))


def _build_comparables_sheet(wb: Workbook, r: AnalysisResult) -> None:
    ws = wb.create_sheet("Comparables")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFG", [28, 14, 14, 12, 12, 12, 20]):
        ws.column_dimensions[col].width = w

    v, f = r.valuation, r.fundamentals

    _style_title_row(ws, 1, f"  {f.company_name}  —  Comparable Company Analysis", 7)

    # Trading comps section
    _style_section(ws, 3, "TRADING COMPS — VALUATION METHODS", 7)
    _style_col_headers(ws, 4, ["Method", "Low", "Mid", "High", "Weight", "vs. Current", "Notes"], 1)

    for i, meth in enumerate(v.methods):
        bg = STRIPE if i % 2 == 0 else WHITE
        row = 5 + i
        _data_cell(ws, row, 1, meth.name, bold=True, bg=bg, align="left")
        _data_cell(ws, row, 2, meth.low or "—", bg=bg)
        _data_cell(ws, row, 3, meth.mid or "—", bg=bg)
        _data_cell(ws, row, 4, meth.high or "—", bg=bg)
        _data_cell(ws, row, 5, f"{meth.weight}%" if meth.weight else "—", bg=bg)
        _data_cell(ws, row, 6, meth.current_pct or "—", bg=bg)
        _data_cell(ws, row, 7, "", bg=bg)

    # DCF assumptions
    dcf = v.dcf_assumptions
    if dcf:
        _style_section(ws, 10, "DCF ASSUMPTIONS", 7)
        _style_col_headers(ws, 11, ["Parameter", "Value"], 1)
        dcf_rows = [
            ("WACC", dcf.wacc or "N/A"),
            ("Terminal Growth Rate", dcf.terminal_growth or "N/A"),
            ("Projection Horizon", f"{dcf.projection_years or 5} years"),
        ]
        for i, (lbl, val) in enumerate(dcf_rows):
            bg = STRIPE if i % 2 == 0 else WHITE
            _data_cell(ws, 12 + i, 1, lbl, bold=True, bg=bg, align="left")
            _data_cell(ws, 12 + i, 2, val, bg=bg)
            ws.merge_cells(f"B{12+i}:G{12+i}")

    # Comparable multiples
    cm = v.comparable_multiples
    if cm:
        _style_section(ws, 16, "MARKET COMPARABLE MULTIPLES", 7)
        _style_col_headers(ws, 17, ["Multiple", "Value"], 1)
        for i, (k, val) in enumerate(cm.items()):
            bg = STRIPE if i % 2 == 0 else WHITE
            _data_cell(ws, 18 + i, 1, str(k).replace("_", "/").upper(), bold=True, bg=bg, align="left")
            _data_cell(ws, 18 + i, 2, str(val), bg=bg)
            ws.merge_cells(f"B{18+i}:G{18+i}")

    # Competitor landscape
    if r.market.main_competitors:
        start = 22
        _style_section(ws, start, "COMPETITIVE LANDSCAPE", 7)
        _style_col_headers(ws, start + 1, ["Company", "Est. Market Share", "Notes"], 1)
        for i, comp in enumerate(r.market.main_competitors):
            bg = STRIPE if i % 2 == 0 else WHITE
            _data_cell(ws, start + 2 + i, 1, comp.name, bold=True, bg=bg, align="left")
            _data_cell(ws, start + 2 + i, 2, comp.market_share or "—", bg=bg)
            _data_cell(ws, start + 2 + i, 3, "", bg=bg)
            ws.merge_cells(f"C{start+2+i}:G{start+2+i}")


def _build_sensitivity_sheet(wb: Workbook, r: AnalysisResult) -> None:
    """WACC (rows) × Terminal Growth Rate (cols) → Equity Value ($M)."""
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False

    v, f, fin = r.valuation, r.fundamentals, r.financials

    for col, w in zip("ABCDEFGH", [20, 14, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    _style_title_row(
        ws, 1,
        f"  {f.company_name}  —  Sensitivity: Equity Value ($M) by WACC × Terminal Growth",
        8,
    )
    ws["A2"].value = "Green = higher equity value   |   Red = lower equity value"
    ws["A2"].font = _font(9, color=MID_GRAY, italic=True)
    ws.merge_cells("A2:H2")

    wacc_range = [0.07, 0.08, 0.09, 0.10, 0.11, 0.12, 0.13]
    tgr_range = [0.01, 0.02, 0.03, 0.04, 0.05]

    # Derive FCF year-5 and sum_pv_fcfs from projection data
    dcf = v.dcf_assumptions
    base_wacc = _parse_pct(dcf.wacc if dcf else None) or 0.10
    ebitda_pct = _parse_pct(fin.ebitda_margin) or 0.15
    cur_rev_m = _to_m(_parse_num(fin.revenue_current)) or 1_000.0
    fcf_conv = 0.70
    net_debt_m = 0.0

    def _calc_equity(wacc: float, tgr: float) -> float:
        """Simple 5-year FCF DCF."""
        rev = cur_rev_m
        sum_pv = 0.0
        fcf_y5 = 0.0
        for n, p in enumerate(range(5), start=1):
            proj = fin.projections[p] if p < len(fin.projections) else None
            rev = _to_m(_parse_num(proj.revenue if proj else None)) or rev * 1.10
            em = _parse_pct(proj.ebitda_margin if proj else None) or ebitda_pct
            fcf = rev * em * fcf_conv
            sum_pv += fcf / (1 + wacc) ** n
            if n == 5:
                fcf_y5 = fcf
        tv = fcf_y5 * (1 + tgr) / max(wacc - tgr, 0.001)
        pv_tv = tv / (1 + wacc) ** 5
        return sum_pv + pv_tv - net_debt_m

    # Header row
    HEADER_ROW = 4
    ws.cell(row=HEADER_ROW, column=1).value = "WACC \\ TGR →"
    ws.cell(row=HEADER_ROW, column=1).font = _font(10, bold=True)
    ws.cell(row=HEADER_ROW, column=1).fill = _fill(DARK_BLUE)
    ws.cell(row=HEADER_ROW, column=1).font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    ws.cell(row=HEADER_ROW, column=1).alignment = _center()

    for j, tgr in enumerate(tgr_range):
        col = j + 2
        c = ws.cell(row=HEADER_ROW, column=col)
        c.value = f"{tgr:.0%}"
        c.fill = _fill(DARK_BLUE)
        c.font = _font(9, bold=True, color=WHITE)
        c.alignment = _center()
        c.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Data rows
    data_start_row = HEADER_ROW + 1
    for i, wacc in enumerate(wacc_range):
        row = data_start_row + i
        # Row label
        c = ws.cell(row=row, column=1)
        c.value = f"{wacc:.0%}"
        c.fill = _fill(DARK_BLUE)
        c.font = _font(9, bold=True, color=WHITE)
        c.alignment = _center()
        c.border = _thin_border()
        ws.row_dimensions[row].height = 20

        for j, tgr in enumerate(tgr_range):
            col = j + 2
            eq = _calc_equity(wacc, tgr)
            cell = ws.cell(row=row, column=col)
            cell.value = round(eq, 0)
            cell.number_format = "#,##0"
            cell.alignment = _center()
            cell.border = _thin_border()
            # Highlight base case
            if abs(wacc - base_wacc) < 0.005 and abs(tgr - (_parse_pct((dcf.terminal_growth if dcf else None) or "3%") or 0.03)) < 0.005:
                cell.fill = _fill(LIGHT_YELLOW)
                cell.font = _font(10, bold=True)

    # Color scale: green (high) → red (low)
    data_area = f"B{data_start_row}:{get_column_letter(1+len(tgr_range))}{data_start_row+len(wacc_range)-1}"
    ws.conditional_formatting.add(
        data_area,
        ColorScaleRule(
            start_type="min", start_color="FFCDD2",
            mid_type="percentile", mid_value=50, mid_color="FFFFFF",
            end_type="max", end_color="C8E6C9",
        ),
    )

    # Notes
    note_row = data_start_row + len(wacc_range) + 2
    ws[f"A{note_row}"].value = (
        "Yellow cell = base case assumptions from DCF Model sheet.  "
        "Values are equity value ($M) = PV(FCFs) + PV(Terminal Value) − Net Debt."
    )
    ws[f"A{note_row}"].font = _font(9, color=MID_GRAY, italic=True)
    ws.merge_cells(f"A{note_row}:H{note_row}")


def _build_financials_sheet(wb: Workbook, r: AnalysisResult) -> None:
    ws = wb.create_sheet("Financials")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDE", [28, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    f, fin = r.fundamentals, r.financials

    _style_title_row(ws, 1, f"  {f.company_name}  —  Financial Summary", 5)

    # Income statement
    if fin.income_statement:
        _style_section(ws, 3, "INCOME STATEMENT", 5)
        period_headers = fin.income_statement[0].values if fin.income_statement else []
        _style_col_headers(ws, 4, ["Line Item"] + list(period_headers), 1)
        for i, row_data in enumerate(fin.income_statement):
            bg = STRIPE if i % 2 == 0 else WHITE
            data_row = 5 + i
            _data_cell(ws, data_row, 1, row_data.line, bold=True, bg=bg, align="left")
            for j, val in enumerate(row_data.values):
                _data_cell(ws, data_row, j + 2, val, bg=bg)

    # Projections
    if fin.projections:
        proj_start = 10
        _style_section(ws, proj_start, "REVENUE PROJECTIONS", 5)
        _style_col_headers(ws, proj_start + 1, ["Year", "Revenue", "Growth", "EBITDA Margin"], 1)
        for i, p in enumerate(fin.projections):
            bg = STRIPE if i % 2 == 0 else WHITE
            row = proj_start + 2 + i
            _data_cell(ws, row, 1, p.year, bold=True, bg=bg)
            _data_cell(ws, row, 2, p.revenue or "N/A", bg=bg)
            _data_cell(ws, row, 3, p.growth or "N/A", bg=bg)
            _data_cell(ws, row, 4, p.ebitda_margin or "N/A", bg=bg)

    # Key metrics
    if fin.key_metrics:
        km_start = 16
        _style_section(ws, km_start, "KEY OPERATING METRICS", 5)
        _style_col_headers(ws, km_start + 1, ["Metric", "Value", "YoY Change"], 1)
        for i, km in enumerate(fin.key_metrics):
            bg = STRIPE if i % 2 == 0 else WHITE
            row = km_start + 2 + i
            _data_cell(ws, row, 1, km.name, bold=True, bg=bg, align="left")
            _data_cell(ws, row, 2, km.value, bg=bg)
            _data_cell(ws, row, 3, km.delta or "—", bg=bg)


# ── 3-Statement model helpers ─────────────────────────────────────────────────

def _revenue_series_m(r: AnalysisResult) -> list[float]:
    """Return 5-year projected revenue series in $M."""
    fin = r.financials
    rev0 = _to_m(_parse_num(fin.revenue_current)) or 0.0
    growth = _parse_pct(fin.revenue_growth) or 0.07
    # Cap wild LLM growth estimates
    growth = max(-0.30, min(growth, 0.50))
    return [rev0 * (1 + growth) ** y for y in range(1, 6)]


def _build_pl_sheet(wb: Workbook, r: AnalysisResult) -> None:
    """P&L — 5-year projected income statement."""
    ws = wb.create_sheet("P&L")
    ws.sheet_view.showGridLines = False

    fin = r.financials
    for col, w in zip("ABCDEFGH", [32, 16, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    _style_title_row(ws, 1, f"  {r.fundamentals.company_name}  —  Projected P&L  ($M, 5-year)", 7)
    ws["A2"].value = "All values in $M  |  Based on stated assumptions"
    ws["A2"].font = _font(9, color=MID_GRAY, italic=True)
    ws.merge_cells("A2:G2")

    years = [f"Y{i}" for i in range(1, 6)]
    headers = ["Line Item"] + years + ["CAGR"]
    _style_col_headers(ws, 4, headers, 1)

    rev_series = _revenue_series_m(r)
    ebitda_m = _parse_pct(fin.ebitda_margin) or 0.18
    gross_m = _parse_pct(fin.gross_margin) or min(ebitda_m + 0.25, 0.80)
    net_m = _parse_pct(fin.net_margin) or (ebitda_m * 0.55)
    da_pct = 0.05  # D&A ~5% of revenue

    def _cagr(s: list[float]) -> str:
        if not s or s[0] <= 0:
            return "—"
        return f"{((s[-1] / s[0]) ** (1 / (len(s) - 1)) - 1):.1%}"

    gross_series = [v * gross_m for v in rev_series]
    ebitda_series = [v * ebitda_m for v in rev_series]
    da_series = [v * da_pct for v in rev_series]
    ebit_series = [e - d for e, d in zip(ebitda_series, da_series)]
    net_series = [v * net_m for v in rev_series]

    sections = [
        ("Revenue", rev_series, LIGHT_BLUE, True),
        ("  COGS", [-v * (1 - gross_m) for v in rev_series], WHITE, False),
        ("Gross Profit", gross_series, STRIPE, True),
        ("  Gross Margin %", [v / r2 if r2 else 0 for v, r2 in zip(gross_series, rev_series)], WHITE, False, "0.0%"),
        ("  OpEx (est.)", [-(v - e) for v, e in zip(gross_series, ebitda_series)], WHITE, False),
        ("EBITDA", ebitda_series, LIGHT_BLUE, True),
        ("  EBITDA Margin %", [v / r2 if r2 else 0 for v, r2 in zip(ebitda_series, rev_series)], WHITE, False, "0.0%"),
        ("  D&A (est.)", [-v for v in da_series], WHITE, False),
        ("EBIT", ebit_series, STRIPE, True),
        ("  Interest (est.)", [0.0] * 5, WHITE, False),
        ("  Tax (25%)", [-v * 0.25 for v in ebit_series], WHITE, False),
        ("Net Income", net_series, LIGHT_BLUE, True),
        ("  Net Margin %", [v / r2 if r2 else 0 for v, r2 in zip(net_series, rev_series)], WHITE, False, "0.0%"),
    ]

    row = 5
    for item in sections:
        label, vals = item[0], item[1]
        bg = item[2]
        bold = item[3]
        fmt = item[4] if len(item) > 4 else "#,##0.0"
        _data_cell(ws, row, 1, label, bold=bold, bg=bg, align="left")
        for ci, v in enumerate(vals, 2):
            _data_cell(ws, row, ci, round(v, 1), bold=bold, bg=bg, fmt=fmt)
        # CAGR for non-margin rows
        if fmt != "0.0%":
            cagr_v = _cagr(vals)
            c = ws.cell(row=row, column=7)
            c.value = cagr_v
            c.font = _font(10, bold=bold, color=MID_GRAY)
            c.alignment = Alignment(horizontal="center")
            c.fill = _fill(bg)
        row += 1

    ws["A2"].value = (
        f"All values $M  |  Revenue growth assumed {(_parse_pct(fin.revenue_growth) or 0.07):.1%}/yr  |"
        f"  EBITDA margin {ebitda_m:.1%}  |  Net margin {net_m:.1%}"
    )


def _build_bs_sheet(wb: Workbook, r: AnalysisResult) -> None:
    """Balance Sheet — simplified 5-year projected."""
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False

    fin = r.financials
    for col, w in zip("ABCDEFGH", [32, 16, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    _style_title_row(ws, 1, f"  {r.fundamentals.company_name}  —  Simplified Balance Sheet  ($M)", 7)
    ws["A2"].value = "Simplified model — actual figures may vary  |  All values in $M"
    ws["A2"].font = _font(9, color=MID_GRAY, italic=True)
    ws.merge_cells("A2:G2")

    years = [f"Y{i}" for i in range(1, 6)]
    headers = ["Line Item"] + years + ["Notes"]
    _style_col_headers(ws, 4, headers, 1)

    rev_series = _revenue_series_m(r)
    ebitda_m = _parse_pct(fin.ebitda_margin) or 0.18
    net_series = [v * (_parse_pct(fin.net_margin) or ebitda_m * 0.55) for v in rev_series]
    da_series = [v * 0.05 for v in rev_series]
    capex_series = [v * 0.04 for v in rev_series]

    d_e = _parse_num(fin.debt_to_equity) or 0.5
    debt0 = (rev_series[0] * ebitda_m * 3) * (d_e / (1 + d_e)) if d_e > 0 else 0
    cash0 = rev_series[0] * 0.08

    # Build cumulative retained earnings
    retained = []
    ret = 0.0
    for ni in net_series:
        ret += ni * 0.7  # assume 30% dividend payout
        retained.append(ret)

    cash_series = [cash0 + sum(net_series[:i + 1]) * 0.3 for i in range(5)]
    ppe_series = [rev_series[i] * 0.25 - sum(capex_series[:i + 1]) + sum(da_series[:i + 1]) for i in range(5)]
    ppe_series = [max(p, 0) for p in ppe_series]
    debt_series = [max(debt0 - sum(capex_series[:i + 1]) * 0.5, 0) for i in range(5)]

    ta_series = [c + r + p + r2 * 0.12 for c, p, r, r2 in zip(cash_series, ppe_series, retained, rev_series)]
    equity_series = [ta - d for ta, d in zip(ta_series, debt_series)]

    sections = [
        ("ASSETS", None, NAVY, True),
        ("  Cash & equivalents", cash_series, LIGHT_BLUE, False),
        ("  Receivables (12% rev)", [v * 0.12 for v in rev_series], WHITE, False),
        ("  Total Current Assets", [c + v * 0.12 for c, v in zip(cash_series, rev_series)], STRIPE, True),
        ("  PP&E (net)", ppe_series, WHITE, False),
        ("Total Assets", ta_series, LIGHT_BLUE, True),
        ("LIABILITIES & EQUITY", None, NAVY, True),
        ("  Payables (8% rev)", [v * 0.08 for v in rev_series], WHITE, False),
        ("  Long-term Debt", debt_series, WHITE, False),
        ("  Total Liabilities", [d + v * 0.08 for d, v in zip(debt_series, rev_series)], STRIPE, True),
        ("  Retained Earnings", retained, WHITE, False),
        ("Total Equity", equity_series, LIGHT_BLUE, True),
    ]

    row = 5
    for item in sections:
        label, vals = item[0], item[1]
        bg = item[2]
        bold = item[3]
        _data_cell(ws, row, 1, label, bold=bold, bg=bg, align="left")
        if vals:
            for ci, v in enumerate(vals, 2):
                _data_cell(ws, row, ci, round(v, 1), bold=bold, bg=bg, fmt="#,##0.0")
        row += 1


def _build_cf_sheet(wb: Workbook, r: AnalysisResult) -> None:
    """Cash Flow Statement — 5-year projected."""
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False

    fin = r.financials
    for col, w in zip("ABCDEFGH", [32, 16, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    _style_title_row(ws, 1, f"  {r.fundamentals.company_name}  —  Cash Flow Statement  ($M)", 7)
    ws["A2"].value = "Simplified model  |  All values in $M"
    ws["A2"].font = _font(9, color=MID_GRAY, italic=True)
    ws.merge_cells("A2:G2")

    years = [f"Y{i}" for i in range(1, 6)]
    _style_col_headers(ws, 4, ["Line Item"] + years + ["Total"], 1)

    rev_series = _revenue_series_m(r)
    ebitda_m = _parse_pct(fin.ebitda_margin) or 0.18
    net_m = _parse_pct(fin.net_margin) or ebitda_m * 0.55
    da_pct = 0.05
    capex_pct = 0.04
    nwc_pct = 0.03  # working capital build ~3% of revenue growth

    net_inc = [v * net_m for v in rev_series]
    da = [v * da_pct for v in rev_series]
    d_rev = [rev_series[i] - (rev_series[i - 1] if i > 0 else rev_series[0] / 1.07) for i in range(5)]
    delta_nwc = [-d * nwc_pct for d in d_rev]
    op_cf = [n + d + w for n, d, w in zip(net_inc, da, delta_nwc)]

    capex = [-v * capex_pct for v in rev_series]
    fcf = [o + c for o, c in zip(op_cf, capex)]

    fin_cf = [0.0] * 5  # simplified: no external financing assumed

    net_cash = [o + c + f for o, c, f in zip(op_cf, capex, fin_cf)]

    sections = [
        ("OPERATING ACTIVITIES", None, NAVY, True),
        ("  Net Income", net_inc, LIGHT_BLUE, False),
        ("  + D&A", da, WHITE, False),
        ("  ± Change in NWC", delta_nwc, WHITE, False),
        ("Operating Cash Flow", op_cf, STRIPE, True),
        ("INVESTING ACTIVITIES", None, NAVY, True),
        ("  Capital Expenditure", capex, WHITE, False),
        ("Free Cash Flow", fcf, LIGHT_BLUE, True),
        ("FINANCING ACTIVITIES", None, NAVY, True),
        ("  Net Debt / Equity Change", fin_cf, WHITE, False),
        ("Net Change in Cash", net_cash, STRIPE, True),
    ]

    row = 5
    for item in sections:
        label, vals = item[0], item[1]
        bg = item[2]
        bold = item[3]
        _data_cell(ws, row, 1, label, bold=bold, bg=bg, align="left")
        if vals:
            for ci, v in enumerate(vals, 2):
                _data_cell(ws, row, ci, round(v, 1), bold=bold, bg=bg, fmt="#,##0.0")
            total = sum(vals)
            c = ws.cell(row=row, column=7)
            c.value = round(total, 1)
            c.font = _font(10, bold=bold, color=DARK_GRAY)
            c.number_format = "#,##0.0"
            c.alignment = Alignment(horizontal="center")
            c.fill = _fill(bg)
        row += 1


# ── Public API ──────────────────────────────────────────────────────────────

def generate_excel(result: AnalysisResult, output_path: str = "analysis.xlsx") -> str:
    """Build and save the full workbook. Returns the output path."""
    wb = Workbook()

    _build_dashboard(wb, result)
    _build_dcf_sheet(wb, result)
    _build_comparables_sheet(wb, result)
    _build_sensitivity_sheet(wb, result)
    _build_financials_sheet(wb, result)
    _build_pl_sheet(wb, result)
    _build_bs_sheet(wb, result)
    _build_cf_sheet(wb, result)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path
